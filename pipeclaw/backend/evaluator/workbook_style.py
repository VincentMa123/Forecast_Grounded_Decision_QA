from __future__ import annotations

from typing import Any, Collection, Sequence

from openpyxl.cell.cell import MergedCell
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


NAVY = "17365D"
BLUE = "1F4E78"
ACCENT = "5B9BD5"
LIGHT_BLUE = "D9EAF7"
PALE_BLUE = "EAF2F8"
WHITE = "FFFFFF"
TEXT = "1F1F1F"
GRID = "D9E2F3"

STATUS_STYLES = {
    "pass": ("E2F0D9", "375623"),
    "not_applicable": ("E7E6E6", "595959"),
    "warning": ("FFF2CC", "7F6000"),
    "pending": ("FFF2CC", "7F6000"),
    "needs_review": ("FCE4D6", "9C0006"),
    "fail": ("F4CCCC", "9C0006"),
}

TAB_COLORS = (
    "4472C4",
    "70AD47",
    "ED7D31",
    "A5A5A5",
    "FFC000",
    "5B9BD5",
)

WIDE_HEADERS = {
    "definition",
    "dispatch_recommendation",
    "evidence_summary",
    "final_answer",
    "issues",
    "path_or_value",
    "purpose",
    "quality_issues",
    "remaining_quality_issues",
    "reviewer_notes",
    "user_input",
}


def polish_workbook(
    workbook: Any,
    illegal_characters: Any,
    *,
    summary_sheets: Collection[str] = (),
) -> None:
    """Apply a consistent, audit-friendly visual system to a workbook."""
    summary_names = set(summary_sheets)
    for index, sheet in enumerate(workbook.worksheets, start=1):
        _sanitize(sheet, illegal_characters)
        sheet.sheet_view.showGridLines = False
        sheet.sheet_view.zoomScale = 85
        sheet.sheet_properties.tabColor = (
            NAVY
            if sheet.title in summary_names
            else TAB_COLORS[(index - 1) % len(TAB_COLORS)]
        )
        sheet.oddFooter.center.text = "Page &P of &N"
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        sheet.page_setup.orientation = "landscape" if sheet.max_column > 8 else "portrait"
        sheet.page_margins.left = 0.25
        sheet.page_margins.right = 0.25
        sheet.page_margins.top = 0.5
        sheet.page_margins.bottom = 0.5
        if sheet.title in summary_names:
            _style_summary(sheet)
        else:
            _style_data_sheet(sheet, index)


def style_chart(
    chart: Any,
    colors: Sequence[str] = ("4472C4", "ED7D31", "70AD47"),
) -> None:
    """Use matching report colors and a restrained chart layout."""
    chart.style = 10
    chart.legend.position = "b"
    chart.height = 7
    chart.width = 12
    for series, color in zip(chart.series, colors):
        series.graphicalProperties.solidFill = color
        series.graphicalProperties.line.solidFill = color


def _sanitize(sheet: Any, illegal_characters: Any) -> None:
    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            if isinstance(cell.value, str):
                cell.value = illegal_characters.sub("", cell.value)[:32767]


def _style_summary(sheet: Any) -> None:
    sheet.freeze_panes = "A2" if sheet.max_row > 1 else None
    sheet.sheet_view.zoomScale = 95
    sheet.row_dimensions[1].height = 34
    sheet.column_dimensions["A"].width = 40
    sheet.column_dimensions["B"].width = 78
    title_fill = PatternFill("solid", fgColor=NAVY)
    for cell in sheet[1]:
        cell.fill = title_fill
        cell.font = Font(name="Arial", size=18, bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="left", vertical="center")
    thin = Side(style="thin", color=GRID)
    for row_number in range(2, sheet.max_row + 1):
        label = sheet.cell(row_number, 1)
        value = sheet.cell(row_number, 2)
        label.font = Font(name="Arial", size=10, bold=True, color=NAVY)
        label.fill = PatternFill(
            "solid",
            fgColor=LIGHT_BLUE if row_number % 2 == 0 else PALE_BLUE,
        )
        label.alignment = Alignment(vertical="center", wrap_text=True)
        value.font = Font(name="Arial", size=10, color=TEXT)
        value.fill = PatternFill(
            "solid",
            fgColor=WHITE if row_number % 2 == 0 else "F8FBFD",
        )
        value.alignment = Alignment(vertical="center", wrap_text=True)
        label.border = Border(bottom=thin)
        value.border = Border(bottom=thin)
        sheet.row_dimensions[row_number].height = 22
        label_text = str(label.value or "").casefold()
        if "rate" in label_text and isinstance(value.value, (int, float)):
            value.number_format = "0.0%"
            fill, color = _rate_style(float(value.value))
            value.fill = PatternFill("solid", fgColor=fill)
            value.font = Font(name="Arial", size=10, bold=True, color=color)
        _apply_status_style(value)
    sheet.auto_filter.ref = None


def _style_data_sheet(sheet: Any, table_index: int) -> None:
    sheet.freeze_panes = "A2" if sheet.max_row > 1 else None
    sheet.print_title_rows = "1:1"
    sheet.row_dimensions[1].height = 34
    header_border = Border(bottom=Side(style="medium", color=ACCENT))
    for cell in sheet[1]:
        if isinstance(cell, MergedCell):
            continue
        cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        cell.border = header_border

    body_border = Border(bottom=Side(style="hair", color=GRID))
    for row_number, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        long_text = False
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            cell.font = Font(name="Arial", size=10, color=TEXT)
            cell.alignment = Alignment(
                horizontal="right" if isinstance(cell.value, (int, float)) else "left",
                vertical="top",
                wrap_text=True,
            )
            cell.border = body_border
            long_text = long_text or (
                isinstance(cell.value, str) and len(cell.value) > 100
            )
            _apply_status_style(cell)
        sheet.row_dimensions[row_number].height = 36 if long_text else 20

    for column_index in range(1, sheet.max_column + 1):
        letter = get_column_letter(column_index)
        header = str(sheet.cell(1, column_index).value or "").casefold()
        sampled = [
            str(sheet.cell(row, column_index).value or "")
            for row in range(1, min(sheet.max_row, 150) + 1)
        ]
        maximum = max(
            (
                max((len(line) for line in value.splitlines()), default=0)
                for value in sampled
            ),
            default=0,
        )
        if header in WIDE_HEADERS or any(
            token in header for token in ("answer", "issue", "summary", "notes")
        ):
            width = 52
        elif any(
            token in header
            for token in ("sample_id", "scenario_id", "session_id", "source")
        ):
            width = min(max(maximum + 2, 18), 34)
        else:
            width = min(max(maximum + 2, 12), 36)
        sheet.column_dimensions[letter].width = width

        if ("rate" in header or "percentage" in header) and sheet.max_row > 1:
            for row in range(2, sheet.max_row + 1):
                sheet.cell(row, column_index).number_format = "0.0%"
            sheet.conditional_formatting.add(
                f"{letter}2:{letter}{sheet.max_row}",
                ColorScaleRule(
                    start_type="num",
                    start_value=0,
                    start_color="F4CCCC",
                    mid_type="num",
                    mid_value=0.8,
                    mid_color="FFF2CC",
                    end_type="num",
                    end_value=1,
                    end_color="E2F0D9",
                ),
            )

    if sheet.max_row > 1 and sheet.max_column > 1:
        table = Table(displayName=f"AuditTable{table_index:02d}", ref=sheet.dimensions)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)
    elif sheet.max_column > 1:
        sheet.auto_filter.ref = sheet.dimensions


def _apply_status_style(cell: Any) -> None:
    if not isinstance(cell.value, str):
        return
    status = cell.value.strip().casefold()
    style = STATUS_STYLES.get(status)
    if style is None:
        return
    fill, color = style
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.font = Font(name="Arial", size=10, bold=True, color=color)
    cell.alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
    )


def _rate_style(value: float) -> tuple[str, str]:
    if value >= 0.9:
        return "E2F0D9", "375623"
    if value >= 0.75:
        return "FFF2CC", "7F6000"
    return "F4CCCC", "9C0006"
