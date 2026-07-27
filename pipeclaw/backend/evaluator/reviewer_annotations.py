from __future__ import annotations

import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, Mapping, Tuple


REVIEW_SHEETS = ("Needs Review", "Manual Spot Check")
REVIEW_COLUMNS = {
    "condition_parsing_review",
    "tool_invocation_review",
    "risk_level_review",
    "dispatch_review",
    "reviewer",
    "reviewer_notes",
    "final_disposition",
}
PENDING_REVIEW_COLUMNS = {
    "condition_parsing_review",
    "tool_invocation_review",
    "risk_level_review",
    "dispatch_review",
    "final_disposition",
}


def export_reviewer_annotations(
    workbook_path: Path,
    output_path: Path,
) -> Dict[str, Any]:
    """Export durable reviewer-entered cells before a report is regenerated."""
    from openpyxl import load_workbook

    workbook_path = Path(workbook_path)
    output_path = Path(output_path)
    existing = load_reviewer_annotations(output_path)
    workbook = load_workbook(workbook_path, read_only=True, data_only=False)
    exported_at = datetime.now(timezone.utc).isoformat()
    found: Dict[Tuple[str, str], Dict[str, Any]] = {}
    for sheet_name in REVIEW_SHEETS:
        if sheet_name not in workbook.sheetnames:
            continue
        sheet = workbook[sheet_name]
        rows = sheet.iter_rows(values_only=True)
        try:
            headers = [str(value or "") for value in next(rows)]
        except StopIteration:
            continue
        for values in rows:
            row = dict(zip(headers, values))
            sample_id = str(row.get("sample_id") or "").strip()
            if not sample_id:
                continue
            review = {
                column: row.get(column)
                for column in headers
                if column in REVIEW_COLUMNS
            }
            if not review:
                continue
            key = (sheet_name, sample_id)
            prior = dict(existing.get(key, {}).get("review") or {})
            for column, value in review.items():
                if _meaningful(value) or not _meaningful(prior.get(column)):
                    prior[column] = value
            found[key] = {
                "sheet": sheet_name,
                "sample_id": sample_id,
                "review": prior,
                "exported_at": exported_at,
                "source_workbook": workbook_path.name,
            }
    workbook.close()
    merged = dict(existing)
    merged.update(found)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8", newline="\n") as handle:
        for key in sorted(merged):
            handle.write(json.dumps(merged[key], ensure_ascii=False) + "\n")
    return {
        "annotation_count": len(merged),
        "exported_from_workbook_count": len(found),
        "output": output_path.as_posix(),
    }


def load_reviewer_annotations(
    path: Path,
) -> Dict[Tuple[str, str], Dict[str, Any]]:
    path = Path(path)
    if not path.is_file():
        return {}
    result: Dict[Tuple[str, str], Dict[str, Any]] = {}
    for raw_line in path.read_text(encoding="utf-8-sig").splitlines():
        if not raw_line.strip():
            continue
        item = json.loads(raw_line)
        sheet = str(item.get("sheet") or "")
        sample_id = str(item.get("sample_id") or "")
        if sheet and sample_id:
            result[(sheet, sample_id)] = item
    return result


def apply_reviewer_annotations(
    workbook: Any,
    annotations: Mapping[Tuple[str, str], Dict[str, Any]],
    *,
    reset_sample_ids: Iterable[str] = (),
) -> None:
    """Restore untouched decisions and reset repaired records to pending review."""
    reset = {str(value) for value in reset_sample_ids}
    for sheet_name in REVIEW_SHEETS:
        if sheet_name not in workbook.sheetnames:
            continue
        sheet = workbook[sheet_name]
        headers = {
            str(cell.value or ""): cell.column
            for cell in sheet[1]
        }
        sample_column = headers.get("sample_id")
        if not sample_column:
            continue
        for row_index in range(2, sheet.max_row + 1):
            sample_id = str(sheet.cell(row_index, sample_column).value or "")
            if not sample_id:
                continue
            if sample_id in reset:
                for column in REVIEW_COLUMNS:
                    column_index = headers.get(column)
                    if not column_index:
                        continue
                    sheet.cell(row_index, column_index).value = (
                        "pending" if column in PENDING_REVIEW_COLUMNS else ""
                    )
                continue
            saved = dict(annotations.get((sheet_name, sample_id), {}).get("review") or {})
            for column, value in saved.items():
                column_index = headers.get(column)
                if column_index:
                    sheet.cell(row_index, column_index).value = value


def load_sample_id_set(path: Path | None) -> set[str]:
    if path is None or not Path(path).is_file():
        return set()
    source = Path(path)
    text = source.read_text(encoding="utf-8-sig")
    if source.suffix.casefold() == ".json":
        value = json.loads(text)
        if isinstance(value, dict):
            value = value.get("sample_ids") or value.get("repaired_sample_ids") or []
        return {str(item) for item in value}
    if source.suffix.casefold() == ".jsonl":
        result = set()
        for line in text.splitlines():
            if not line.strip():
                continue
            item = json.loads(line)
            if isinstance(item, str):
                result.add(item)
            elif item.get("sample_id"):
                result.add(str(item["sample_id"]))
            else:
                result.update(str(value) for value in item.get("sample_ids") or [])
        return result
    return {line.strip() for line in text.splitlines() if line.strip()}


def _meaningful(value: Any) -> bool:
    return value not in (None, "", "pending")
