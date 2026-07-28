from __future__ import annotations

import hashlib
import json
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Mapping, Optional, Sequence

from .teacher_quality import (
    numeric_claim_values,
    numeric_claims_are_grounded,
    numeric_grounding_evidence,
    tool_output_failed,
)

from .tool_evidence import attach_tool_arguments

TASK1_REQUIRED_FIELDS = (
    "sample_id",
    "scenario_id",
    "scenario_type",
    "state_before",
    "recent_turns",
    "user_input",
    "parsed_task",
    "tool_calls",
    "tool_outputs",
    "prediction_summary",
    "constraint_check",
    "evidence",
    "risk_level",
    "manual_intervention_label",
    "dispatch_recommendation",
    "final_answer",
    "quality_flag",
)

EXPECTED_TYPES = {
    "sample_id": str,
    "scenario_id": str,
    "scenario_type": str,
    "state_before": dict,
    "recent_turns": list,
    "user_input": str,
    "parsed_task": dict,
    "tool_calls": list,
    "tool_outputs": list,
    "prediction_summary": dict,
    "constraint_check": dict,
    "evidence": dict,
    "final_answer": str,
    "quality_flag": str,
}

ENTIRELY_SAFE_CLAIM = re.compile(
    r"完全安全|无任何风险|没有任何风险|所有(?:校核|规则|约束)均?通过|各项(?:校核|规则|约束)均?通过"
    r"|\b(?:entirely|completely|fully)\s+safe\b"
    r"|\ball\s+(?:requested\s+)?(?:checks|constraints|rules)\s+pass(?:ed)?\b"
    r"|\bno\s+(?:operational\s+)?risk\b",
    re.IGNORECASE,
)
REDUCE_UPSTREAM_INJECTION = re.compile(
    r"(?:减少|降低|下调|削减).{0,24}(?:上游|气源).{0,16}(?:注气|供气|供给|流量)"
    r"|(?:上游|气源).{0,16}(?:注气|供气|供给|流量).{0,24}(?:减少|降低|下调|削减)"
    r"|\b(?:reduce|decrease|lower|cut)\b.{0,40}\b(?:upstream\s+)?(?:injection|supply|inflow)\b",
    re.IGNORECASE,
)
RAISE_COMPRESSOR_LOAD = re.compile(
    r"(?:提高|增加|上调).{0,24}压缩机.{0,12}负荷"
    r"|压缩机.{0,12}负荷.{0,24}(?:提高|增加|上调)"
    r"|\b(?:raise|increase|boost)\b.{0,40}\bcompressor\s+load\b",
    re.IGNORECASE,
)


@dataclass(frozen=True)
class Task1VerificationConfig:
    manual_sample_rate: float = 0.25
    manual_sample_seed: str = "task1-quality-v1"

    def __post_init__(self) -> None:
        if not 0.20 <= self.manual_sample_rate <= 0.30:
            raise ValueError("manual_sample_rate must be between 0.20 and 0.30.")


class Task1QualityVerifier:
    """Run the Task 1.8 checks and produce the Task 1.9 audit artifacts."""

    def __init__(self, config: Optional[Task1VerificationConfig] = None) -> None:
        self.config = config or Task1VerificationConfig()

    def evaluate(
        self,
        record: Dict[str, Any],
        native_result: Mapping[str, Any],
    ) -> Dict[str, Any]:
        schema = self._schema_check(record)
        numerical = self._numerical_check(record)
        rule = self._rule_check(record)
        dispatch = self._dispatch_check(record)
        checks = {
            "schema": schema,
            "numerical_consistency": numerical,
            "rule_consistency": rule,
            "dispatch_consistency": dispatch,
        }
        failed = [name for name, value in checks.items() if value["status"] == "fail"]
        task1_flag = (
            "pass"
            if native_result.get("quality_flag") == "pass" and not failed
            else "needs_review"
        )
        return {
            "sample_id": record.get("sample_id"),
            "scenario_id": record.get("scenario_id"),
            "dataset_source": record.get("dataset_source"),
            "scenario_type": record.get("scenario_type"),
            "split": record.get("split"),
            "native_quality_flag": native_result.get("quality_flag"),
            "native_quality_score": native_result.get("quality_score"),
            "native_failed_checks": native_result.get("failed_checks") or [],
            "native_quality_issues": native_result.get("quality_issues") or [],
            "task1_quality_flag": task1_flag,
            "task1_failed_checks": failed,
            "evidence_item_count": self.evidence_item_count(record.get("evidence") or {}),
            "final_answer_chars": len(str(record.get("final_answer") or "")),
            "checks": checks,
        }

    @staticmethod
    def _schema_check(record: Dict[str, Any]) -> Dict[str, Any]:
        missing = [field for field in TASK1_REQUIRED_FIELDS if field not in record]
        invalid_types = [
            field
            for field, expected in EXPECTED_TYPES.items()
            if field in record and not isinstance(record[field], expected)
        ]
        nullable_type_issues = [
            field
            for field in ("risk_level", "manual_intervention_label", "dispatch_recommendation")
            if field in record and record[field] is not None and not isinstance(record[field], str)
        ]
        issues = [f"missing:{field}" for field in missing]
        issues.extend(f"invalid_type:{field}" for field in invalid_types + nullable_type_issues)
        return {
            "status": "pass" if not issues else "fail",
            "issues": issues,
        }

    @staticmethod
    def _numerical_check(record: Dict[str, Any]) -> Dict[str, Any]:
        answer = str(record.get("final_answer") or "")
        question = str(record.get("user_input") or "")
        evidence = numeric_grounding_evidence(record)
        claims = numeric_claim_values(answer)
        grounded = numeric_claims_are_grounded(answer, question, evidence)
        return {
            "status": "pass" if grounded else "fail",
            "claimed_numeric_value_count": len(claims),
            "issues": [] if grounded else ["unsupported_numerical_claim"],
        }

    @staticmethod
    def _rule_check(record: Dict[str, Any]) -> Dict[str, Any]:
        constraint = dict(record.get("constraint_check") or {})
        if not constraint:
            return {"status": "not_applicable", "issues": []}

        issues: List[str] = []
        expected_risk = constraint.get("risk_level")
        expected_intervention = constraint.get("human_intervention_label")
        if expected_risk is not None and record.get("risk_level") != expected_risk:
            issues.append("risk_level_disagrees_with_constraint_check")
        if (
            expected_intervention is not None
            and record.get("manual_intervention_label") != expected_intervention
        ):
            issues.append("intervention_label_disagrees_with_constraint_check")

        category_status = dict(constraint.get("category_status") or {})
        rule_status = dict(constraint.get("rule_status") or {})
        nonpass = any(value in {"warning", "fail"} for value in category_status.values())
        nonpass = nonpass or any(value in {"warning", "fail"} for value in rule_status.values())
        if nonpass and ENTIRELY_SAFE_CLAIM.search(str(record.get("final_answer") or "")):
            issues.append("final_answer_claims_entirely_safe_despite_nonpass_rule")

        pressure_fail = category_status.get("pressure") == "fail" or any(
            str(flag).startswith("pressure_violation")
            for flag in constraint.get("triggered_flags") or []
        )
        if pressure_fail and record.get("risk_level") == "low":
            issues.append("pressure_violation_cannot_have_low_risk")
        if pressure_fail and record.get("manual_intervention_label") == "no_intervention":
            issues.append("pressure_violation_cannot_require_no_intervention")
        return {
            "status": "pass" if not issues else "fail",
            "issues": issues,
            "overall_constraint_status": constraint.get("overall_status"),
        }

    @staticmethod
    def _dispatch_check(record: Dict[str, Any]) -> Dict[str, Any]:
        constraint = dict(record.get("constraint_check") or {})
        if not constraint:
            return {"status": "not_applicable", "issues": []}

        issues: List[str] = []
        category_status = dict(constraint.get("category_status") or {})
        rule_status = dict(constraint.get("rule_status") or {})
        flags = {str(value) for value in constraint.get("triggered_flags") or []}
        pressure_fail = (
            category_status.get("pressure") == "fail"
            or any(value.startswith("pressure_violation") for value in flags)
            or rule_status.get("node_pressure_operating_window") == "fail"
        )
        compressor_overload = (
            "compressor_overload" in flags
            or rule_status.get("compressor_load_limit") == "fail"
        )
        dispatch = "\n".join(
            value
            for value in (
                str(record.get("dispatch_recommendation") or "").strip(),
                str(record.get("final_answer") or "").strip(),
            )
            if value
        )
        if pressure_fail and REDUCE_UPSTREAM_INJECTION.search(dispatch):
            issues.append("pressure_violation_recommends_reducing_upstream_injection")
        if compressor_overload and RAISE_COMPRESSOR_LOAD.search(dispatch):
            issues.append("compressor_overload_recommends_raising_compressor_load")

        expected = constraint.get("dispatch_recommendation")
        actual = record.get("dispatch_recommendation")
        if expected and actual and str(expected).strip() != str(actual).strip():
            issues.append("dispatch_recommendation_disagrees_with_constraint_check")
        return {
            "status": "pass" if not issues else "fail",
            "issues": issues,
            "pressure_failure_present": pressure_fail,
            "compressor_overload_present": compressor_overload,
        }

    @staticmethod
    def evidence_item_count(value: Any) -> int:
        if value is None or value == "" or value == [] or value == {}:
            return 0
        if isinstance(value, list):
            return sum(max(1, Task1QualityVerifier.evidence_item_count(item)) for item in value)
        if isinstance(value, dict):
            return sum(Task1QualityVerifier.evidence_item_count(item) for item in value.values())
        return 1

    @staticmethod
    def manual_review_queue(
        records: Sequence[Dict[str, Any]],
        evaluations: Sequence[Dict[str, Any]],
    ) -> List[Dict[str, Any]]:
        """Return every record still requiring native or Task 1 review."""
        by_id = {str(item.get("sample_id")): item for item in evaluations}
        selected = [
            record
            for record in records
            if (
                (by_id.get(str(record.get("sample_id"))) or {}).get("native_quality_flag")
                != "pass"
                or (by_id.get(str(record.get("sample_id"))) or {}).get("task1_quality_flag")
                != "pass"
            )
        ]
        return sorted(selected, key=lambda item: str(item.get("sample_id")))

    @staticmethod
    def manual_evidence_summary(record: Dict[str, Any]) -> str:
        """Return compact rule and watch-variable evidence for human review."""
        constraint = dict(record.get("constraint_check") or {})
        constraint_items = [constraint]
        constraint_items.extend(
            dict(item) for item in constraint.get("candidate_forecasts") or []
        )
        findings = []
        for item in constraint_items:
            for finding in item.get("priority_findings") or []:
                finding = dict(finding)
                findings.append({
                    "name": finding.get("name"),
                    "status": finding.get("status"),
                    "affected_variables": list(finding.get("affected_variables") or [])[:5],
                })

        evidence = dict(record.get("evidence") or {})
        evidence_items = [evidence]
        evidence_items.extend(dict(item) for item in evidence.get("candidate_forecasts") or [])
        watch = []
        for item in evidence_items:
            for variable in item.get("top_watch_variables") or []:
                variable = dict(variable)
                watch.append({
                    "variable": variable.get("variable"),
                    "mean_prediction": variable.get("mean_prediction"),
                    "mean_abs_delta_vs_observed": variable.get("mean_abs_delta_vs_observed"),
                })
        return json.dumps(
            {
                "priority_findings": findings[:5],
                "top_watch_variables": watch[:5],
            },
            ensure_ascii=False,
            separators=(",", ":"),
        )
    def manual_sample(
        self,
        records: Sequence[Dict[str, Any]],
        evaluations: Sequence[Dict[str, Any]],
    ) -> List[Dict[str, Any]]:
        by_id = {str(item.get("sample_id")): item for item in evaluations}
        groups: Dict[tuple[str, str, str], List[Dict[str, Any]]] = defaultdict(list)
        for record in records:
            result = by_id.get(str(record.get("sample_id"))) or {}
            key = (
                str(record.get("dataset_source") or "unknown"),
                str(record.get("scenario_type") or "unknown"),
                str(result.get("task1_quality_flag") or "unknown"),
            )
            groups[key].append(record)

        selected: List[Dict[str, Any]] = []
        for group_records in groups.values():
            count = max(1, round(len(group_records) * self.config.manual_sample_rate))
            ranked = sorted(
                group_records,
                key=lambda item: hashlib.sha256(
                    f"{self.config.manual_sample_seed}:{item.get('sample_id')}".encode("utf-8")
                ).hexdigest(),
            )
            selected.extend(ranked[:count])
        return sorted(selected, key=lambda item: str(item.get("sample_id")))

    @staticmethod
    def statistics(
        records: Sequence[Dict[str, Any]],
        evaluations: Sequence[Dict[str, Any]],
    ) -> Dict[str, Any]:
        by_id = {str(item.get("sample_id")): item for item in evaluations}
        constraint_types = Counter()
        task_types = Counter()
        evidence_counts = []
        for record in records:
            parsed = dict(record.get("parsed_task") or {})
            if record.get("scenario_type") == "pipeformer":
                task_types[str(parsed.get("task_type") or "unspecified")] += 1
            for category in (record.get("constraint_check") or {}).get("requested_categories") or []:
                constraint_types[str(category)] += 1
            evidence_counts.append(Task1QualityVerifier.evidence_item_count(record.get("evidence") or {}))

        task1_flags = Counter(
            str(by_id.get(str(record.get("sample_id")), {}).get("task1_quality_flag") or "unknown")
            for record in records
        )
        native_flags = Counter(
            str(
                by_id.get(str(record.get("sample_id")), {}).get("native_quality_flag")
                or "unknown"
            )
            for record in records
        )
        total = len(records)
        return {
            "total_sample_count": total,
            "task_type_distribution": dict(sorted(task_types.items())),
            "scenario_type_distribution": dict(sorted(Counter(str(r.get("scenario_type") or "unknown") for r in records).items())),
            "dataset_source_distribution": dict(sorted(Counter(str(r.get("dataset_source") or "unknown") for r in records).items())),
            "constraint_type_distribution": dict(sorted(constraint_types.items())),
            "risk_level_distribution": dict(sorted(Counter(str(r.get("risk_level") or "not_applicable") for r in records).items())),
            "human_intervention_distribution": dict(sorted(Counter(str(r.get("manual_intervention_label") or "not_applicable") for r in records).items())),
            "native_quality_distribution": dict(sorted(native_flags.items())),
            "task1_quality_distribution": dict(sorted(task1_flags.items())),
            "average_evidence_item_count": round(sum(evidence_counts) / total, 6) if total else 0.0,
            "native_quality_pass_rate": round(native_flags.get("pass", 0) / total, 6) if total else 0.0,
            "task1_quality_pass_rate": round(task1_flags.get("pass", 0) / total, 6) if total else 0.0,
        }

    @staticmethod
    def schema() -> Dict[str, Any]:
        nullable_string = {"type": ["string", "null"]}
        return {
            "$schema": "https://json-schema.org/draft/2020-12/schema",
            "$id": "teacher_trace_schema.json",
            "title": "PipeClaw Task 1 Teacher Trace Record",
            "type": "object",
            "required": list(TASK1_REQUIRED_FIELDS),
            "properties": {
                "sample_id": {"type": "string", "minLength": 1},
                "scenario_id": {"type": "string", "minLength": 1},
                "scenario_type": {"type": "string", "minLength": 1},
                "state_before": {
                    "type": "object",
                    "required": ["schema_version", "scope", "provenance"],
                    "properties": {
                        "schema_version": {
                            "const": "verified_decision_state_v1"
                        },
                        "scope": {"type": "object"},
                        "verified_evidence": {"type": "object"},
                        "registry_variables": {
                            "type": "object",
                            "required": [
                                "context_only",
                                "search_call_ids",
                                "returned_ids",
                            ],
                        },
                        "candidates": {"type": "array"},
                        "decision_policy": {"type": "object"},
                        "applied_disturbances": {"type": "array"},
                        "unresolved_inputs": {"type": "array"},
                        "provenance": {"type": "object"},
                    },
                    "additionalProperties": False,
                },
                "recent_turns": {
                    "type": "array",
                    "maxItems": 2,
                    "items": {"type": "object"},
                },
                "user_input": {"type": "string"},
                "parsed_task": {"type": "object"},
                "tool_calls": {"type": "array", "items": {"type": "object"}},
                "tool_outputs": {"type": "array", "items": {"type": "object"}},
                "prediction_summary": {"type": "object"},
                "constraint_check": {"type": "object"},
                "evidence": {"type": "object"},
                "risk_level": {
                    "type": ["string", "null"],
                    "enum": ["low", "medium", "high", "unknown", None],
                },
                "manual_intervention_label": {
                    "type": ["string", "null"],
                    "enum": [
                        "no_intervention",
                        "monitoring_only",
                        "operator_attention_required",
                        "immediate_intervention_required",
                        "unknown",
                        None,
                    ],
                },
                "dispatch_recommendation": nullable_string,
                "final_answer": {"type": "string"},
                "quality_flag": {"type": "string", "enum": ["pass", "needs_review"]},
            },
            "additionalProperties": True,
        }

    @staticmethod
    def write_schema(path: Path) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(
            json.dumps(Task1QualityVerifier.schema(), ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )

    @staticmethod
    def write_audit_splits(
        output_dir: Path,
        records: Sequence[Dict[str, Any]],
        compact_split_dir: Optional[Path] = None,
        quality_sample_ids: Optional[set[str]] = None,
    ) -> Dict[str, int]:
        output_dir.mkdir(parents=True, exist_ok=True)
        allowed_ids: Optional[set[str]] = None
        if compact_split_dir and compact_split_dir.is_dir():
            allowed_ids = set()
            for split in ("train", "valid", "test"):
                source = compact_split_dir / f"teacher_trace_{split}.jsonl"
                if source.is_file():
                    for line in source.read_text(encoding="utf-8-sig").splitlines():
                        if line.strip():
                            allowed_ids.add(str(json.loads(line).get("sample_id") or ""))

        counts: Dict[str, int] = {}
        for split in ("train", "valid", "test"):
            selected = [
                {key: value for key, value in record.items() if key != "split"}
                for record in records
                if record.get("split") == split
                and (
                    (
                        quality_sample_ids is not None
                        and str(record.get("sample_id") or "") in quality_sample_ids
                    )
                    or (
                        quality_sample_ids is None
                        and record.get("quality_flag") == "pass"
                    )
                )
                and not record.get("sft_exclusion_reason")
                and all(
                    item.get("quality_flag") == "pass"
                    for item in record.get("conversation_context") or []
                )
                and (allowed_ids is None or str(record.get("sample_id") or "") in allowed_ids)
            ]
            target = output_dir / f"teacher_trace_{split}.jsonl"
            with target.open("w", encoding="utf-8", newline="\n") as handle:
                for record in selected:
                    handle.write(json.dumps(record, ensure_ascii=False, separators=(",", ":")) + "\n")
            counts[split] = len(selected)
        return counts

    def write_report(
        self,
        path: Path,
        records: Sequence[Dict[str, Any]],
        evaluations: Sequence[Dict[str, Any]],
        statistics: Mapping[str, Any],
        manual_records: Sequence[Dict[str, Any]],
        source_path: Path,
        artifacts: Mapping[str, Any],
        reviewer_annotations: Optional[Mapping[Any, Dict[str, Any]]] = None,
        reset_review_sample_ids: Optional[Sequence[str]] = None,
    ) -> None:
        try:
            from openpyxl import Workbook
            from openpyxl.chart import BarChart, Reference
            from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
            from openpyxl.styles import Font, PatternFill

            from .workbook_style import polish_workbook, style_chart
        except ImportError as exc:
            raise RuntimeError(
                "openpyxl is required for teacher_trace_quality_report.xlsx; "
                "install pipeclaw/backend/requirements.txt."
            ) from exc

        path.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        summary = workbook.active
        summary.title = "Summary"
        check_summary = workbook.create_sheet("Check Summary")
        checks_sheet = workbook.create_sheet("Record Checks")
        constraint_coverage = workbook.create_sheet("Constraint Coverage")
        rule_outcomes = workbook.create_sheet("Rule Outcomes")
        distributions = workbook.create_sheet("Distributions")
        issues_sheet = workbook.create_sheet("Quality Issues")
        needs_review = workbook.create_sheet("Needs Review")
        manual = workbook.create_sheet("Manual Spot Check")
        dataset_coverage = workbook.create_sheet("Dataset Coverage")
        deliverables = workbook.create_sheet("Deliverables")
        notes = workbook.create_sheet("Method Notes")
        workbook.properties.title = "Task 1 Teacher Trace Quality Report"
        workbook.properties.subject = "Section 1.8 quality control and Section 1.9 statistics"

        by_id = {str(item.get("sample_id")): item for item in evaluations}
        generated_at = datetime.now(timezone.utc).isoformat()
        check_counts = {
            name: Counter(item["checks"][name]["status"] for item in evaluations)
            for name in (
                "schema",
                "numerical_consistency",
                "rule_consistency",
                "dispatch_consistency",
            )
        }
        summary_rows = [
            ("Task 1 Teacher Trace Quality Report", None),
            ("Generated at (UTC)", generated_at),
            ("Teacher trace source", source_path.as_posix()),
            ("Total sample count", statistics["total_sample_count"]),
            ("Native quality pass count", statistics["native_quality_distribution"].get("pass", 0)),
            ("Native quality pass rate", statistics["native_quality_pass_rate"]),
            ("Task 1 verification pass count", statistics["task1_quality_distribution"].get("pass", 0)),
            ("Task 1 verification pass rate", statistics["task1_quality_pass_rate"]),
            ("Schema check pass count", check_counts["schema"].get("pass", 0)),
            ("Numerical consistency pass count", check_counts["numerical_consistency"].get("pass", 0)),
            ("Numerical consistency failure count", check_counts["numerical_consistency"].get("fail", 0)),
            ("Applicable rule-consistency checks", sum(check_counts["rule_consistency"].values()) - check_counts["rule_consistency"].get("not_applicable", 0)),
            ("Rule-consistency failure count", check_counts["rule_consistency"].get("fail", 0)),
            ("Applicable dispatch-consistency checks", sum(check_counts["dispatch_consistency"].values()) - check_counts["dispatch_consistency"].get("not_applicable", 0)),
            ("Dispatch-consistency failure count", check_counts["dispatch_consistency"].get("fail", 0)),
            ("Average evidence item count", statistics["average_evidence_item_count"]),
            ("Manual spot-check sample count", len(manual_records)),
            ("Manual spot-check sample rate", len(manual_records) / len(records) if records else 0.0),
            ("Manual review status", "pending_human_signoff"),
        ]
        for row in summary_rows:
            summary.append(row)
        summary["A1"].font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
        summary["A1"].fill = PatternFill("solid", fgColor="17365D")
        summary.merge_cells("A1:B1")
        for row in range(2, summary.max_row + 1):
            if "rate" in str(summary.cell(row=row, column=1).value).lower():
                summary.cell(row=row, column=2).number_format = "0.0%"
        summary.column_dimensions["A"].width = 38
        summary.column_dimensions["B"].width = 95

        check_summary.append([
            "check",
            "pass",
            "fail",
            "not_applicable",
            "applicable_count",
            "applicable_pass_rate",
            "purpose",
        ])
        check_purposes = {
            "schema": "Required Task 1.7 fields and JSON container types.",
            "numerical_consistency": "Final-answer numeric claims are present in trusted evidence.",
            "rule_consistency": "Risk and intervention conclusions agree with constraint_check.",
            "dispatch_consistency": "Recommendations obey pressure and compressor safety priorities.",
        }
        for name, counts in check_counts.items():
            applicable = counts.get("pass", 0) + counts.get("fail", 0)
            check_summary.append([
                name,
                counts.get("pass", 0),
                counts.get("fail", 0),
                counts.get("not_applicable", 0),
                applicable,
                counts.get("pass", 0) / applicable if applicable else None,
                check_purposes[name],
            ])
            check_summary.cell(check_summary.max_row, 6).number_format = "0.0%"

        check_headers = [
            "sample_id",
            "dataset_source",
            "scenario_id",
            "session_id",
            "turn_id",
            "scenario_type",
            "split",
            "task_type",
            "tool_call_count",
            "successful_tool_output_count",
            "failed_tool_output_count",
            "native_profile",
            "native_quality_flag",
            "native_quality_score",
            "task1_quality_flag",
            "evidence_item_count",
            "final_answer_chars",
            "schema_check",
            "numerical_consistency",
            "rule_consistency",
            "dispatch_consistency",
            "overall_constraint_status",
            "risk_level",
            "manual_intervention_label",
            "failed_checks",
            "issues",
        ]
        checks_sheet.append(check_headers)
        for record in records:
            result = by_id[str(record.get("sample_id"))]
            task_checks = result["checks"]
            check_issues = [
                issue
                for check in task_checks.values()
                for issue in check.get("issues") or []
            ]
            outputs = attach_tool_arguments(
                record.get("tool_outputs") or [], record.get("tool_calls") or []
            )
            checks_sheet.append([
                record.get("sample_id"),
                record.get("dataset_source"),
                record.get("scenario_id"),
                record.get("session_id"),
                record.get("turn_id"),
                record.get("scenario_type"),
                record.get("split"),
                (record.get("parsed_task") or {}).get("task_type") or record.get("scenario_type"),
                len(record.get("tool_calls") or []),
                sum(not tool_output_failed(item) for item in outputs),
                sum(tool_output_failed(item) for item in outputs),
                result.get("native_profile"),
                result.get("native_quality_flag"),
                result.get("native_quality_score"),
                result.get("task1_quality_flag"),
                result.get("evidence_item_count"),
                result.get("final_answer_chars"),
                task_checks["schema"]["status"],
                task_checks["numerical_consistency"]["status"],
                task_checks["rule_consistency"]["status"],
                task_checks["dispatch_consistency"]["status"],
                (record.get("constraint_check") or {}).get("overall_status"),
                record.get("risk_level"),
                record.get("manual_intervention_label"),
                ", ".join(result.get("task1_failed_checks") or []),
                ", ".join((result.get("native_quality_issues") or []) + check_issues),
            ])

        requested_categories = Counter()
        category_counts: Dict[str, Counter[str]] = defaultdict(Counter)
        category_scenarios: Dict[str, set[str]] = defaultdict(set)
        rule_counts: Dict[str, Counter[str]] = defaultdict(Counter)
        for record in records:
            constraint = dict(record.get("constraint_check") or {})
            category_status = dict(constraint.get("category_status") or {})
            for category in constraint.get("requested_categories") or []:
                category = str(category)
                requested_categories[category] += 1
                category_counts[category][str(category_status.get(category) or "missing")] += 1
                category_scenarios[category].add(str(record.get("scenario_id") or ""))
            for rule, status in (constraint.get("rule_status") or {}).items():
                rule_counts[str(rule)][str(status or "missing")] += 1

        constraint_coverage.append([
            "constraint_category",
            "requested_count",
            "evaluated_count",
            "pass",
            "warning",
            "fail",
            "not_evaluated_or_missing",
            "evaluation_coverage_rate",
            "non_pass_rate",
            "scenario_count",
        ])
        for category in sorted(requested_categories):
            counts = category_counts[category]
            requested = requested_categories[category]
            evaluated = counts.get("pass", 0) + counts.get("warning", 0) + counts.get("fail", 0)
            nonpass = counts.get("warning", 0) + counts.get("fail", 0)
            constraint_coverage.append([
                category,
                requested,
                evaluated,
                counts.get("pass", 0),
                counts.get("warning", 0),
                counts.get("fail", 0),
                requested - evaluated,
                evaluated / requested if requested else None,
                nonpass / evaluated if evaluated else None,
                len(category_scenarios[category]),
            ])
            constraint_coverage.cell(constraint_coverage.max_row, 8).number_format = "0.0%"
            constraint_coverage.cell(constraint_coverage.max_row, 9).number_format = "0.0%"

        rule_outcomes.append([
            "rule_id",
            "evaluated_count",
            "pass",
            "warning",
            "fail",
            "not_evaluated_or_missing",
            "pass_rate",
            "non_pass_rate",
        ])
        for rule in sorted(rule_counts):
            counts = rule_counts[rule]
            evaluated = counts.get("pass", 0) + counts.get("warning", 0) + counts.get("fail", 0)
            nonpass = counts.get("warning", 0) + counts.get("fail", 0)
            other = sum(counts.values()) - evaluated
            rule_outcomes.append([
                rule,
                evaluated,
                counts.get("pass", 0),
                counts.get("warning", 0),
                counts.get("fail", 0),
                other,
                counts.get("pass", 0) / evaluated if evaluated else None,
                nonpass / evaluated if evaluated else None,
            ])
            rule_outcomes.cell(rule_outcomes.max_row, 7).number_format = "0.0%"
            rule_outcomes.cell(rule_outcomes.max_row, 8).number_format = "0.0%"

        coverage_groups: Dict[tuple[str, str, str], List[Dict[str, Any]]] = defaultdict(list)
        for record in records:
            coverage_groups[(
                str(record.get("dataset_source") or "unknown"),
                str(record.get("scenario_type") or "unknown"),
                str(record.get("split") or "unknown"),
            )].append(record)
        dataset_coverage.append([
            "dataset_source",
            "scenario_type",
            "split",
            "sample_count",
            "scenario_count",
            "session_count",
            "task1_pass_count",
            "task1_pass_rate",
            "average_evidence_items",
            "average_final_answer_chars",
            "average_tool_calls",
        ])
        for key, group in sorted(coverage_groups.items()):
            group_results = [by_id[str(item.get("sample_id"))] for item in group]
            passed = sum(item.get("task1_quality_flag") == "pass" for item in group_results)
            dataset_coverage.append([
                *key,
                len(group),
                len({str(item.get("scenario_id") or "") for item in group}),
                len({str(item.get("session_id") or "") for item in group}),
                passed,
                passed / len(group),
                sum(item.get("evidence_item_count", 0) for item in group_results) / len(group),
                sum(item.get("final_answer_chars", 0) for item in group_results) / len(group),
                sum(len(item.get("tool_calls") or []) for item in group) / len(group),
            ])
            dataset_coverage.cell(dataset_coverage.max_row, 8).number_format = "0.0%"

        needs_review.append([
            "sample_id",
            "dataset_source",
            "scenario_id",
            "session_id",
            "turn_id",
            "scenario_type",
            "split",
            "native_quality_score",
            "native_failed_checks",
            "task1_failed_checks",
            "quality_issues",
            "user_input",
            "final_answer",
            "reviewer_notes",
            "final_disposition",
        ])
        for record in records:
            result = by_id[str(record.get("sample_id"))]
            if result.get("task1_quality_flag") == "pass":
                continue
            needs_review.append([
                record.get("sample_id"),
                record.get("dataset_source"),
                record.get("scenario_id"),
                record.get("session_id"),
                record.get("turn_id"),
                record.get("scenario_type"),
                record.get("split"),
                result.get("native_quality_score"),
                ", ".join(result.get("native_failed_checks") or []),
                ", ".join(result.get("task1_failed_checks") or []),
                ", ".join(result.get("native_quality_issues") or []),
                record.get("user_input"),
                record.get("final_answer"),
                "",
                "pending",
            ])

        distributions.append(["dimension", "value", "count", "percentage"])
        distribution_keys = (
            "dataset_source_distribution",
            "task_type_distribution",
            "scenario_type_distribution",
            "constraint_type_distribution",
            "risk_level_distribution",
            "human_intervention_distribution",
            "native_quality_distribution",
            "task1_quality_distribution",
        )
        for dimension in distribution_keys:
            values = statistics[dimension]
            denominator = sum(values.values()) or 1
            for value, count in values.items():
                distributions.append([dimension, value, count, count / denominator])
                distributions.cell(distributions.max_row, 4).number_format = "0.0%"

        issue_counts = Counter()
        issue_samples: Dict[str, List[str]] = defaultdict(list)
        issue_datasets: Dict[str, set[str]] = defaultdict(set)
        issue_scenarios: Dict[str, set[str]] = defaultdict(set)
        for record in records:
            result = by_id[str(record.get("sample_id"))]
            record_issues = {str(issue) for issue in result.get("native_quality_issues") or []}
            record_issues.update(
                str(issue)
                for check in result["checks"].values()
                for issue in check.get("issues") or []
            )
            issue_counts.update(record_issues)
            for issue in record_issues:
                issue_samples[issue].append(str(record.get("sample_id") or ""))
                issue_datasets[issue].add(str(record.get("dataset_source") or "unknown"))
                issue_scenarios[issue].add(str(record.get("scenario_id") or "unknown"))
        issues_sheet.append([
            "issue",
            "record_occurrences",
            "percentage_of_dataset",
            "dataset_sources",
            "scenario_count",
            "affected_sample_ids",
        ])
        for issue, count in sorted(issue_counts.items()):
            issues_sheet.append([
                issue,
                count,
                count / len(records) if records else 0.0,
                ", ".join(sorted(issue_datasets[issue])),
                len(issue_scenarios[issue]),
                ", ".join(issue_samples[issue]),
            ])
            issues_sheet.cell(issues_sheet.max_row, 3).number_format = "0.0%"

        manual.append([
            "sample_id",
            "dataset_source",
            "scenario_id",
            "session_id",
            "turn_id",
            "scenario_type",
            "split",
            "automated_quality",
            "native_quality_score",
            "remaining_quality_issues",
            "schema_check",
            "numerical_consistency",
            "rule_consistency",
            "dispatch_consistency",
            "user_input",
            "final_answer",
            "parsed_task_summary",
            "tool_names",
            "constraint_status_summary",
            "evidence_summary",
            "risk_level",
            "intervention_label",
            "dispatch_recommendation",
            "condition_parsing_review",
            "tool_invocation_review",
            "risk_level_review",
            "dispatch_review",
            "reviewer",
            "reviewer_notes",
            "final_disposition",
        ])
        for record in manual_records:
            result = by_id[str(record.get("sample_id"))]
            task_checks = result["checks"]
            manual.append([
                record.get("sample_id"),
                record.get("dataset_source"),
                record.get("scenario_id"),
                record.get("session_id"),
                record.get("turn_id"),
                record.get("scenario_type"),
                record.get("split"),
                result.get("task1_quality_flag"),
                result.get("native_quality_score"),
                ", ".join(sorted(
                    {str(issue) for issue in result.get("native_quality_issues") or []}
                    | {
                        str(issue)
                        for check in task_checks.values()
                        for issue in check.get("issues") or []
                    }
                    | {f"native_check:{name}" for name in result.get("native_failed_checks") or []}
                    | {f"task1_check:{name}" for name in result.get("task1_failed_checks") or []}
                )),
                task_checks["schema"]["status"],
                task_checks["numerical_consistency"]["status"],
                task_checks["rule_consistency"]["status"],
                task_checks["dispatch_consistency"]["status"],
                record.get("user_input"),
                record.get("final_answer"),
                json.dumps(record.get("parsed_task") or {}, ensure_ascii=False, separators=(",", ":")),
                ", ".join(str(item.get("name") or "") for item in record.get("tool_calls") or []),
                json.dumps((record.get("constraint_check") or {}).get("category_status") or {}, ensure_ascii=False, separators=(",", ":")),
                self.manual_evidence_summary(record),
                record.get("risk_level"),
                record.get("manual_intervention_label"),
                record.get("dispatch_recommendation"),
                "pending",
                "pending",
                "pending",
                "pending",
                "",
                "",
                "pending",
            ])

        deliverables.append(["artifact", "path_or_value"])
        for name, value in artifacts.items():
            deliverables.append([name, json.dumps(value, ensure_ascii=False) if isinstance(value, (dict, list)) else value])

        notes.append(["topic", "definition"])
        notes.append(["Schema check", "Required Task 1.7 fields must exist and use the expected JSON container types."])
        notes.append(["Numerical consistency", "Numbers in final_answer must be grounded in the question or trusted parsed-task, forecast, constraint, evidence, conversation, or successful-tool data."])
        notes.append(["Rule consistency", "Top-level risk/intervention fields must agree with constraint_check; a non-pass rule cannot be summarized as entirely safe."])
        notes.append(["Dispatch consistency", "Pressure violations cannot recommend reducing upstream injection, and compressor overload cannot recommend raising compressor load."])
        notes.append(["Evidence item count", "Recursive count of non-empty scalar/list evidence items in the top-level evidence field."])
        notes.append(["Manual spot check", "Repair runs queue every remaining needs_review record; evaluation-only runs use deterministic stratified sampling. Pending columns require human engineering sign-off."])
        notes.append(["Workbook formulas", "This report is a fixed verification snapshot and intentionally contains no formulas or external links."])

        polish_workbook(
            workbook,
            ILLEGAL_CHARACTERS_RE,
            summary_sheets={"Summary"},
        )
        if check_summary.max_row > 1:
            chart = BarChart()
            chart.type = "bar"
            chart.grouping = "stacked"
            chart.overlap = 100
            chart.title = "Quality Checks: Pass vs Fail"
            chart.y_axis.title = "check"
            chart.x_axis.title = "records"
            chart.add_data(
                Reference(
                    check_summary,
                    min_col=2,
                    max_col=3,
                    min_row=1,
                    max_row=check_summary.max_row,
                ),
                titles_from_data=True,
            )
            chart.set_categories(
                Reference(
                    check_summary,
                    min_col=1,
                    min_row=2,
                    max_row=check_summary.max_row,
                )
            )
            style_chart(chart, ("70AD47", "C00000"))
            check_summary.add_chart(chart, "I2")

        if reviewer_annotations:
            from .reviewer_annotations import apply_reviewer_annotations

            apply_reviewer_annotations(
                workbook,
                reviewer_annotations,
                reset_sample_ids=reset_review_sample_ids or (),
            )
        workbook.save(path)

    @staticmethod
    def verify_workbook(path: Path) -> Dict[str, Any]:
        from openpyxl import load_workbook

        workbook = load_workbook(path, data_only=False, read_only=True)
        sheet_count = len(workbook.sheetnames)
        error_tokens = {"#REF!", "#DIV/0!", "#VALUE!", "#N/A", "#NAME?"}
        formulas = 0
        errors: List[str] = []
        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows():
                for cell in row:
                    value = cell.value
                    if isinstance(value, str) and value.startswith("="):
                        formulas += 1
                    if isinstance(value, str) and any(token in value for token in error_tokens):
                        errors.append(f"{worksheet.title}!{cell.coordinate}")
        workbook.close()
        return {
            "sheet_count": sheet_count,
            "formula_count": formulas,
            "error_count": len(errors),
            "error_cells": errors[:20],
        }
