from __future__ import annotations

import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Mapping, Optional, Sequence

from .teacher_trace_audit import TeacherTraceQualityAuditor


class TeacherTraceQualityReportWriter:
    """Write teacher-trace schemas, audit splits, and quality workbooks."""

    def __init__(self, auditor: TeacherTraceQualityAuditor) -> None:
        self.auditor = auditor

    def write_schema(self, path: Path) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(
            json.dumps(self.auditor.schema(), ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )

    @staticmethod
    def write_audit_splits(
        output_dir: Path,
        records: Sequence[Dict[str, Any]],
        compact_split_dir: Optional[Path] = None,
        quality_sample_ids: Optional[set[str]] = None,
    ) -> Dict[str, int]:
        output_dir.mkdir(parents=True, exist_ok=True)
        allowed_ids: Optional[set[str]] = None
        if compact_split_dir and compact_split_dir.is_dir():
            allowed_ids = set()
            for split in ("train", "valid", "test"):
                source = compact_split_dir / f"teacher_trace_{split}.jsonl"
                if source.is_file():
                    for line in source.read_text(encoding="utf-8-sig").splitlines():
                        if line.strip():
                            allowed_ids.add(str(json.loads(line).get("sample_id") or ""))

        counts: Dict[str, int] = {}
        for split in ("train", "valid", "test"):
            selected = [
                {key: value for key, value in record.items() if key != "split"}
                for record in records
                if record.get("split") == split
                and (
                    (
                        quality_sample_ids is not None
                        and str(record.get("sample_id") or "") in quality_sample_ids
                    )
                    or (
                        quality_sample_ids is None
                        and record.get("quality_flag") == "pass"
                    )
                )
                and not record.get("sft_exclusion_reason")
                and all(
                    item.get("quality_flag") == "pass"
                    for item in record.get("conversation_context") or []
                )
                and (allowed_ids is None or str(record.get("sample_id") or "") in allowed_ids)
            ]
            target = output_dir / f"teacher_trace_{split}.jsonl"
            with target.open("w", encoding="utf-8", newline="\n") as handle:
                for record in selected:
                    handle.write(json.dumps(record, ensure_ascii=False, separators=(",", ":")) + "\n")
            counts[split] = len(selected)
        return counts

    def write_report(
        self,
        path: Path,
        facts: Mapping[str, Any],
        manual_records: Sequence[Dict[str, Any]],
        source_path: Path,
        artifacts: Mapping[str, Any],
        reviewer_annotations: Optional[Mapping[Any, Dict[str, Any]]] = None,
        reset_review_sample_ids: Optional[Sequence[str]] = None,
    ) -> None:
        try:
            from openpyxl import Workbook
            from openpyxl.chart import BarChart, Reference
            from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
            from openpyxl.styles import Font, PatternFill

            from .workbook_style import polish_workbook, style_chart
        except ImportError as exc:
            raise RuntimeError(
                "openpyxl is required for teacher_trace_quality_report.xlsx; "
                "install pipeclaw/backend/requirements.txt."
            ) from exc

        path.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        summary = workbook.active
        summary.title = "Summary"
        check_summary = workbook.create_sheet("Check Summary")
        checks_sheet = workbook.create_sheet("Record Checks")
        constraint_coverage = workbook.create_sheet("Constraint Coverage")
        rule_outcomes = workbook.create_sheet("Rule Outcomes")
        distributions = workbook.create_sheet("Distributions")
        issues_sheet = workbook.create_sheet("Quality Issues")
        needs_review = workbook.create_sheet("Needs Review")
        manual = workbook.create_sheet("Manual Spot Check")
        dataset_coverage = workbook.create_sheet("Dataset Coverage")
        deliverables = workbook.create_sheet("Deliverables")
        notes = workbook.create_sheet("Method Notes")
        workbook.properties.title = "Task 1 Teacher Trace Quality Report"
        workbook.properties.subject = "Section 1.8 quality control and Section 1.9 statistics"

        generated_at = datetime.now(timezone.utc).isoformat()
        summary_rows = [
            ("Task 1 Teacher Trace Quality Report", None),
            ("Generated at (UTC)", generated_at),
            ("Teacher trace source", source_path.as_posix()),
            *facts["quality_summary_rows"],
            ("Manual spot-check sample count", len(manual_records)),
            (
                "Manual spot-check sample rate",
                len(manual_records) / facts["statistics"]["total_sample_count"]
                if facts["statistics"]["total_sample_count"]
                else 0.0,
            ),
            ("Manual review status", "pending_human_signoff"),
        ]
        for row in summary_rows:
            summary.append(row)
        summary["A1"].font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
        summary["A1"].fill = PatternFill("solid", fgColor="17365D")
        summary.merge_cells("A1:B1")
        for row in range(2, summary.max_row + 1):
            if "rate" in str(summary.cell(row=row, column=1).value).lower():
                summary.cell(row=row, column=2).number_format = "0.0%"
        summary.column_dimensions["A"].width = 38
        summary.column_dimensions["B"].width = 95

        check_summary.append([
            "check",
            "pass",
            "fail",
            "not_applicable",
            "applicable_count",
            "applicable_pass_rate",
            "purpose",
        ])
        for row in facts["check_summary_rows"]:
            check_summary.append(row)
            check_summary.cell(check_summary.max_row, 6).number_format = "0.0%"

        check_headers = [
            "sample_id",
            "dataset_source",
            "scenario_id",
            "session_id",
            "turn_id",
            "scenario_type",
            "split",
            "task_type",
            "tool_call_count",
            "successful_tool_output_count",
            "failed_tool_output_count",
            "native_profile",
            "native_quality_flag",
            "native_quality_score",
            "task1_quality_flag",
            "evidence_item_count",
            "final_answer_chars",
            "schema_check",
            "numerical_consistency",
            "rule_consistency",
            "dispatch_consistency",
            "overall_constraint_status",
            "risk_level",
            "manual_intervention_label",
            "failed_checks",
            "issues",
        ]
        checks_sheet.append(check_headers)
        for row in facts["record_check_rows"]:
            checks_sheet.append(row)

        constraint_coverage.append([
            "constraint_category",
            "requested_count",
            "evaluated_count",
            "pass",
            "warning",
            "fail",
            "not_evaluated_or_missing",
            "evaluation_coverage_rate",
            "non_pass_rate",
            "scenario_count",
        ])
        for row in facts["quality_category_rows"]:
            constraint_coverage.append(row)
            constraint_coverage.cell(constraint_coverage.max_row, 8).number_format = "0.0%"
            constraint_coverage.cell(constraint_coverage.max_row, 9).number_format = "0.0%"

        rule_outcomes.append([
            "rule_id",
            "evaluated_count",
            "pass",
            "warning",
            "fail",
            "not_evaluated_or_missing",
            "pass_rate",
            "non_pass_rate",
        ])
        for row in facts["rule_rows"]:
            rule_outcomes.append(row)
            rule_outcomes.cell(rule_outcomes.max_row, 7).number_format = "0.0%"
            rule_outcomes.cell(rule_outcomes.max_row, 8).number_format = "0.0%"

        dataset_coverage.append([
            "dataset_source",
            "scenario_type",
            "split",
            "sample_count",
            "scenario_count",
            "session_count",
            "task1_pass_count",
            "task1_pass_rate",
            "average_evidence_items",
            "average_final_answer_chars",
            "average_tool_calls",
        ])
        for row in facts["coverage_rows"]:
            dataset_coverage.append(row)
            dataset_coverage.cell(dataset_coverage.max_row, 8).number_format = "0.0%"

        needs_review.append([
            "sample_id",
            "dataset_source",
            "scenario_id",
            "session_id",
            "turn_id",
            "scenario_type",
            "split",
            "native_quality_score",
            "native_failed_checks",
            "task1_failed_checks",
            "quality_issues",
            "user_input",
            "final_answer",
            "reviewer_notes",
            "final_disposition",
        ])
        for row in facts["needs_review_rows"]:
            needs_review.append(row)

        distributions.append(["dimension", "value", "count", "percentage"])
        for row in facts["distribution_rows"]:
            distributions.append(row)
            distributions.cell(distributions.max_row, 4).number_format = "0.0%"

        issues_sheet.append([
            "issue",
            "record_occurrences",
            "percentage_of_dataset",
            "dataset_sources",
            "scenario_count",
            "affected_sample_ids",
        ])
        for row in facts["issue_rows"]:
            issues_sheet.append(row)
            issues_sheet.cell(issues_sheet.max_row, 3).number_format = "0.0%"

        manual.append([
            "sample_id",
            "dataset_source",
            "scenario_id",
            "session_id",
            "turn_id",
            "scenario_type",
            "split",
            "automated_quality",
            "native_quality_score",
            "remaining_quality_issues",
            "schema_check",
            "numerical_consistency",
            "rule_consistency",
            "dispatch_consistency",
            "user_input",
            "final_answer",
            "parsed_task_summary",
            "tool_names",
            "constraint_status_summary",
            "evidence_summary",
            "risk_level",
            "intervention_label",
            "dispatch_recommendation",
            "condition_parsing_review",
            "tool_invocation_review",
            "risk_level_review",
            "dispatch_review",
            "reviewer",
            "reviewer_notes",
            "final_disposition",
        ])
        for record in manual_records:
            manual.append(facts["manual_rows_by_id"][str(record.get("sample_id") or "")])

        deliverables.append(["artifact", "path_or_value"])
        for name, value in artifacts.items():
            deliverables.append([name, json.dumps(value, ensure_ascii=False) if isinstance(value, (dict, list)) else value])

        notes.append(["topic", "definition"])
        notes.append(["Schema check", "Required Task 1.7 fields must exist and use the expected JSON container types."])
        notes.append(["Numerical consistency", "Numbers in final_answer must be grounded in the question or trusted parsed-task, forecast, constraint, evidence, conversation, or successful-tool data."])
        notes.append(["Rule consistency", "Top-level risk/intervention fields must agree with constraint_check; a non-pass rule cannot be summarized as entirely safe."])
        notes.append(["Dispatch consistency", "Pressure violations cannot recommend reducing upstream injection, and compressor overload cannot recommend raising compressor load."])
        notes.append(["Evidence item count", "Recursive count of non-empty scalar/list evidence items in the top-level evidence field."])
        notes.append(["Manual spot check", "Repair runs queue every remaining needs_review record; evaluation-only runs use deterministic stratified sampling. Pending columns require human engineering sign-off."])
        notes.append(["Workbook formulas", "This report is a fixed verification snapshot and intentionally contains no formulas or external links."])

        polish_workbook(
            workbook,
            ILLEGAL_CHARACTERS_RE,
            summary_sheets={"Summary"},
        )
        if check_summary.max_row > 1:
            chart = BarChart()
            chart.type = "bar"
            chart.grouping = "stacked"
            chart.overlap = 100
            chart.title = "Quality Checks: Pass vs Fail"
            chart.y_axis.title = "check"
            chart.x_axis.title = "records"
            chart.add_data(
                Reference(
                    check_summary,
                    min_col=2,
                    max_col=3,
                    min_row=1,
                    max_row=check_summary.max_row,
                ),
                titles_from_data=True,
            )
            chart.set_categories(
                Reference(
                    check_summary,
                    min_col=1,
                    min_row=2,
                    max_row=check_summary.max_row,
                )
            )
            style_chart(chart, ("70AD47", "C00000"))
            check_summary.add_chart(chart, "I2")

        if reviewer_annotations:
            from .reviewer_annotations import apply_reviewer_annotations

            apply_reviewer_annotations(
                workbook,
                reviewer_annotations,
                reset_sample_ids=reset_review_sample_ids or (),
            )
        workbook.save(path)

    @staticmethod
    def verify_workbook(path: Path) -> Dict[str, Any]:
        from openpyxl import load_workbook

        workbook = load_workbook(path, data_only=False, read_only=True)
        sheet_count = len(workbook.sheetnames)
        error_tokens = {"#REF!", "#DIV/0!", "#VALUE!", "#N/A", "#NAME?"}
        formulas = 0
        errors: List[str] = []
        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows():
                for cell in row:
                    value = cell.value
                    if isinstance(value, str) and value.startswith("="):
                        formulas += 1
                    if isinstance(value, str) and any(token in value for token in error_tokens):
                        errors.append(f"{worksheet.title}!{cell.coordinate}")
        workbook.close()
        return {
            "sheet_count": sheet_count,
            "formula_count": formulas,
            "error_count": len(errors),
            "error_cells": errors[:20],
        }
