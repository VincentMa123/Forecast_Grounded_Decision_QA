from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Mapping

class Task1StatisticsWorkbook:
    """Create the standalone Task 1.9 data-statistics workbook."""

    @staticmethod
    def write(
        path: Path,
        facts: Mapping[str, Any],
        source_path: Path,
    ) -> None:
        from openpyxl import Workbook
        from openpyxl.chart import BarChart, PieChart, Reference
        from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
        from openpyxl.styles import Font, PatternFill

        from .workbook_style import polish_workbook, style_chart

        path.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        overview = workbook.active
        overview.title = "Overview"
        source_sheet = workbook.create_sheet("Dataset Sources")
        scenario_sheet = workbook.create_sheet("Scenario Types")
        task_sheet = workbook.create_sheet("Task Types")
        constraint_sheet = workbook.create_sheet("Constraint Types")
        outcome_sheet = workbook.create_sheet("Constraint Outcomes")
        rule_sheet = workbook.create_sheet("Rule Outcomes")
        risk_sheet = workbook.create_sheet("Risk Levels")
        intervention_sheet = workbook.create_sheet("Intervention Labels")
        cross_sheet = workbook.create_sheet("Risk x Intervention")
        split_sheet = workbook.create_sheet("Split Statistics")
        evidence_sheet = workbook.create_sheet("Evidence Statistics")
        quality_sheet = workbook.create_sheet("Quality Statistics")
        notes = workbook.create_sheet("Method Notes")
        workbook.properties.title = "Task 1 Teacher Trace Data Statistics"
        workbook.properties.subject = "Task 1.9 data statistics tables"

        overview_rows = [
            ("Task 1 Teacher Trace Data Statistics", None),
            ("Generated at (UTC)", datetime.now(timezone.utc).isoformat()),
            ("Teacher trace source", source_path.as_posix()),
            *facts["statistics_overview_rows"],
        ]
        for row in overview_rows:
            overview.append(row)
        overview["A1"].font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
        overview["A1"].fill = PatternFill("solid", fgColor="17365D")
        overview.merge_cells("A1:B1")
        for row in range(2, overview.max_row + 1):
            if "rate" in str(overview.cell(row, 1).value).lower():
                overview.cell(row, 2).number_format = "0.0%"

        source_sheet.append(["metric", *facts["source_columns"]])
        for row in facts["source_rows"]:
            source_sheet.append(row)
            if row[0] in {"percentage", "quality_pass_rate"}:
                for cell in source_sheet[source_sheet.max_row][1:]:
                    cell.number_format = "0.0%"

        for sheet, distribution, label in (
            (scenario_sheet, "scenario_type_distribution", "scenario_type"),
            (task_sheet, "task_type_distribution", "pipeformer_task_type"),
            (constraint_sheet, "constraint_type_distribution", "constraint_type"),
            (risk_sheet, "risk_level_distribution", "risk_level"),
            (
                intervention_sheet,
                "human_intervention_distribution",
                "intervention_label",
            ),
        ):
            Task1StatisticsWorkbook._distribution_sheet(
                sheet,
                facts["distribution_tables"][distribution],
                label,
            )

        outcome_sheet.append([
            "constraint_category", "requested", "evaluated", "pass", "warning", "fail",
            "not_evaluated_or_missing", "coverage_rate", "pass_rate", "non_pass_rate",
        ])
        for row in facts["statistics_category_rows"]:
            outcome_sheet.append(row)
            for column in (8, 9, 10):
                outcome_sheet.cell(outcome_sheet.max_row, column).number_format = "0.0%"

        rule_sheet.append([
            "rule_id", "evaluated", "pass", "warning", "fail", "not_evaluated_or_missing",
            "pass_rate", "non_pass_rate",
        ])
        for row in facts["rule_rows"]:
            rule_sheet.append(row)
            rule_sheet.cell(rule_sheet.max_row, 7).number_format = "0.0%"
            rule_sheet.cell(rule_sheet.max_row, 8).number_format = "0.0%"

        cross_sheet.append(["risk_level", *facts["risk_intervention_columns"], "row_total"])
        for row in facts["risk_intervention_rows"]:
            cross_sheet.append(row)

        split_sheet.append([
            "split", "master_samples", "scenario_count", "session_count", "quality_pass",
            "needs_review", "sft_eligible", "sft_eligible_rate", "average_evidence_items",
            "average_answer_chars",
        ])
        for row in facts["split_rows"]:
            split_sheet.append(row)
            split_sheet.cell(split_sheet.max_row, 8).number_format = "0.0%"

        evidence_sheet.append([
            "dataset_source", "scenario_type", "sample_count", "zero_evidence_samples",
            "minimum", "p25", "median", "mean", "p75", "p90", "p95", "maximum",
        ])
        for row in facts["evidence_rows"]:
            evidence_sheet.append(row)

        quality_sheet.append([
            "quality_dimension", "pass", "fail_or_needs_review", "not_applicable",
            "applicable_count", "applicable_pass_rate",
        ])
        for row in facts["quality_rows"]:
            quality_sheet.append(row)
            quality_sheet.cell(quality_sheet.max_row, 6).number_format = "0.0%"

        notes.append(["topic", "definition"])
        notes.append(["Purpose", "Standalone Task 1.9 data-statistics tables for the master teacher-trace dataset."])
        notes.append(["SFT eligible", "Task 1 pass, no explicit SFT exclusion, and all preceding conversation turns pass quality."])
        notes.append(["Evidence items", "Recursive count of non-empty scalar/list items in the top-level evidence object."])
        notes.append(["Scenario types", "Scenario-family distribution across all teacher-trace records."])
        notes.append(["Task types", "PipeFormer task-subtype distribution across PipeFormer scenario records only; unspecified means no subtype was stored."])
        notes.append(["Constraint requested", "A category listed in constraint_check.requested_categories."])
        notes.append(["Constraint evaluated", "Requested category with pass, warning, or fail in category_status."])
        notes.append(["Percentiles", "Nearest-rank interpolation over per-record evidence-item counts."])
        notes.append(["Workbook formulas", "Fixed verification snapshot; no formulas or external links are used."])

        polish_workbook(
            workbook,
            ILLEGAL_CHARACTERS_RE,
            summary_sheets={"Overview"},
        )
        source_sheet.freeze_panes = "B2"
        for cell in source_sheet["A"][1:]:
            cell.font = Font(name="Arial", size=10, bold=True, color="1F1F1F")
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
        if source_sheet.max_column > 1:
            chart = BarChart()
            chart.type = "bar"
            chart.style = 10
            chart.title = "Samples by Dataset Source"
            chart.y_axis.title = "dataset_source"
            chart.x_axis.title = "sample_count"
            chart.add_data(
                Reference(source_sheet, min_col=1, max_col=source_sheet.max_column, min_row=2, max_row=2),
                titles_from_data=True,
                from_rows=True,
            )
            chart.set_categories(Reference(source_sheet, min_col=2, max_col=source_sheet.max_column, min_row=1))
            style_chart(chart, ("4472C4",))
            source_sheet.add_chart(chart, "F2")
        for sheet, title in (
            (scenario_sheet, "Scenario-Type Distribution"),
            (task_sheet, "PipeFormer Task-Type Distribution"),
            (constraint_sheet, "Constraint-Type Distribution"),
        ):
            Task1StatisticsWorkbook._add_bar_chart(
                sheet, BarChart, Reference, title, 2, "F2"
            )
        if risk_sheet.max_row > 1:
            chart = PieChart()
            chart.title = "Risk-Level Distribution"
            chart.add_data(Reference(risk_sheet, min_col=2, min_row=1, max_row=risk_sheet.max_row), titles_from_data=True)
            chart.set_categories(Reference(risk_sheet, min_col=1, min_row=2, max_row=risk_sheet.max_row))
            style_chart(
                chart,
                ("4472C4", "70AD47", "FFC000", "ED7D31", "C00000"),
            )
            risk_sheet.add_chart(chart, "F2")
        workbook.save(path)

    @staticmethod
    def _distribution_sheet(sheet: Any, rows: Any, label: str) -> None:
        sheet.append([label, "count", "percentage"])
        for row in rows:
            sheet.append(row)
            sheet.cell(sheet.max_row, 3).number_format = "0.0%"

    @staticmethod
    def _add_bar_chart(sheet: Any, chart_cls: Any, reference_cls: Any, title: str, value_column: int, anchor: str) -> None:
        if sheet.max_row < 2:
            return
        chart = chart_cls()
        chart.type = "bar"
        chart.style = 10
        chart.title = title
        chart.y_axis.title = sheet.cell(1, 1).value
        chart.x_axis.title = sheet.cell(1, value_column).value
        chart.add_data(reference_cls(sheet, min_col=value_column, min_row=1, max_row=sheet.max_row), titles_from_data=True)
        chart.set_categories(reference_cls(sheet, min_col=1, min_row=2, max_row=sheet.max_row))
        chart.height = 7
        chart.width = 12
        from .workbook_style import style_chart

        style_chart(chart, ("4472C4",))
        sheet.add_chart(chart, anchor)
