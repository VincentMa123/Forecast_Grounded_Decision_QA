from __future__ import annotations

import json
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

from reporting.workbook_style import polish_workbook
from scripts.deterministic_repairs import DETERMINISTIC_REPAIR_SAMPLE_IDS
from scripts.repair_teacher_trace import REGENERATION_TARGETS


BACKEND_ROOT = Path(__file__).resolve().parents[1]
GENERATED_ROOT = BACKEND_ROOT / "generated_teacher_traces"
MASTER_TRACE = GENERATED_ROOT / "teacher_trace.json"
QUALITY_REPORT = GENERATED_ROOT / "task1_deliverables" / "teacher_trace_quality_report.xlsx"
AUDIT_EVALUATION = GENERATED_ROOT / "pipeformer_audit_eval" / "quality_evaluation.jsonl"
OUTPUT = GENERATED_ROOT / "task1_deliverables" / "pipeformer_additional_audit.xlsx"
PIPEFORMER_DATASETS = {
    "Pipeline_Full_Life_Cycle_Test_Dataset-v4",
    "Pipeline_Full_Life_Cycle_Test_Dataset-v7",
}


ADDITIONAL_REGENERATION = {
    ("Pipeline_Full_Life_Cycle_Test_Dataset-v4", "scenario_pipeformer_dispatch_003"): (
        "reclassify_existing_failure",
        "A stored forecast call conflicted with the B_302:FR disturbance. The later answer repair does not remove the failed tool trajectory.",
        "scenario_pipeformer_dispatch_003_session_001::turn_001",
    ),
    ("Pipeline_Full_Life_Cycle_Test_Dataset-v4", "scenario_pipeformer_dispatch_005"): (
        "reclassify_existing_failure",
        "A stored forecast call conflicted with the E_018:SNQ disturbance. Regeneration is cleaner than retaining the failed call plus answer-only repairs.",
        "scenario_pipeformer_dispatch_005_session_001::turn_001",
    ),
    ("Pipeline_Full_Life_Cycle_Test_Dataset-v4", "scenario_pipeformer_dispatch_007"): (
        "new_uncovered_scenario",
        "The workbook marked the sampled answer pass, but the trajectory contains a failed B_001:FR forecast call caused by a disturbance/action conflict.",
        "scenario_pipeformer_dispatch_007_session_001::turn_001",
    ),
    ("Pipeline_Full_Life_Cycle_Test_Dataset-v4", "scenario_pipeformer_dispatch_009"): (
        "new_uncovered_scenario",
        "The trajectory contains a failed B_120:FR forecast call caused by applying an action to the same disturbed variable in the opposite direction.",
        "scenario_pipeformer_dispatch_009_session_001::turn_001",
    ),
    ("Pipeline_Full_Life_Cycle_Test_Dataset-v4", "scenario_pipeformer_dispatch_012"): (
        "new_uncovered_scenario",
        "Binary C_008:ST is modeled as a 20% disturbance instead of a discrete 0/1 state, so the forecasts were produced under an invalid state transition.",
        "scenario_pipeformer_dispatch_012_session_001::turn_001",
    ),
    ("Pipeline_Full_Life_Cycle_Test_Dataset-v4", "scenario_pipeformer_dispatch_013"): (
        "new_uncovered_scenario",
        "The sampled answer was accepted, but the stored trajectory includes a failed R_001:SPD forecast call caused by a disturbance/action conflict.",
        "scenario_pipeformer_dispatch_013_session_001::turn_001",
    ),
    ("Pipeline_Full_Life_Cycle_Test_Dataset-v4", "scenario_pipeformer_dispatch_014"): (
        "new_uncovered_scenario",
        "The stored trajectory includes a failed R_004:SPD forecast call before the valid alternatives were evaluated.",
        "scenario_pipeformer_dispatch_014_session_001::turn_001",
    ),
    ("Pipeline_Full_Life_Cycle_Test_Dataset-v4", "scenario_pipeformer_dispatch_015"): (
        "reclassify_existing_failure",
        "R_006:ST is discrete, but non-restore candidates encode its shutdown as -100%. Answer-only repair cannot correct the forecast inputs and verification.",
        "scenario_pipeformer_dispatch_015_session_001::turn_001",
    ),
    ("Pipeline_Full_Life_Cycle_Test_Dataset-v4", "scenario_pipeformer_dispatch_019"): (
        "new_uncovered_scenario",
        "The stored trajectory includes a failed T_005:SNQ forecast call caused by a disturbance/action conflict.",
        "scenario_pipeformer_dispatch_019_session_001::turn_001",
    ),
    ("Pipeline_Full_Life_Cycle_Test_Dataset-v4", "scenario_pipeformer_prediction_015"): (
        "new_uncovered_scenario",
        "The first forecast call failed registry normalization, and R_006:ST closure is incorrectly audited as a 100% percentage adjustment instead of a binary setpoint.",
        "scenario_pipeformer_prediction_015_session_001::turn_001",
    ),
    ("Pipeline_Full_Life_Cycle_Test_Dataset-v7", "scenario_pipeformer_dispatch_004"): (
        "new_uncovered_scenario",
        "The trajectory contains a failed C_023:SP_out forecast call; the answer also does not substantively rank compressor-load reduction despite that being the primary objective.",
        "scenario_pipeformer_dispatch_004_session_001::turn_001",
    ),
    ("Pipeline_Full_Life_Cycle_Test_Dataset-v7", "scenario_pipeformer_dispatch_008"): (
        "new_uncovered_scenario",
        "Two candidate forecast calls failed because B_329:FR was simultaneously used as the disturbance and an opposing control action.",
        "scenario_pipeformer_dispatch_008_session_001::turn_001",
    ),
    ("Pipeline_Full_Life_Cycle_Test_Dataset-v7", "scenario_pipeformer_dispatch_012"): (
        "new_uncovered_scenario",
        "Binary C_017:ST shutdown is encoded as a -100% disturbance rather than a discrete setpoint, invalidating the boundary-state contract.",
        "scenario_pipeformer_dispatch_012_session_001::turn_001",
    ),
    ("Pipeline_Full_Life_Cycle_Test_Dataset-v7", "scenario_pipeformer_dispatch_015"): (
        "reclassify_existing_failure",
        "R_005:ST shutdown is encoded as a -100% disturbance. The existing answer repair fixes ranking text but not the invalid forecast inputs.",
        "scenario_pipeformer_dispatch_015_session_001::turn_001",
    ),
    ("Pipeline_Full_Life_Cycle_Test_Dataset-v7", "scenario_pipeformer_prediction_015"): (
        "new_uncovered_scenario",
        "Although setpoint 0 is applied, verification still treats closure as a 100% adjustment and incorrectly labels the user-specified closure as an LLM assumption.",
        "scenario_pipeformer_prediction_015_session_001::turn_001",
    ),
    ("Pipeline_Full_Life_Cycle_Test_Dataset-v7", "scenario_pipeformer_prediction_016"): (
        "new_uncovered_scenario",
        "The user explicitly supplied +18%, but the structured prediction marks direction and magnitude as an LLM assumption. This is a provenance/parsing failure in the tool result.",
        "scenario_pipeformer_prediction_016_session_001::turn_001",
    ),
}


ANSWER_REPAIRS = {
    "Pipeline_Full_Life_Cycle_Test_Dataset-v4:scenario_pipeformer_prediction_016_session_001::turn_002": (
        "Remove the claim that the explicit E_082:SNQ +18% input was an LLM assumption. The underlying forecast and verification are otherwise usable."
    ),
    "Pipeline_Full_Life_Cycle_Test_Dataset-v7:scenario_pipeformer_dispatch_009_session_001::turn_001": (
        "Provide a complete ranking for candidates 1 and 2 instead of leaving them unordered; keep candidate 3 first because it is the only zero-warning option."
    ),
    "Pipeline_Full_Life_Cycle_Test_Dataset-v7:scenario_pipeformer_dispatch_013_session_001::turn_001": (
        "Honor the pressure-first objective: candidate 2 is the pressure leader. Rank all three candidates and use energy only after the pressure criterion."
    ),
    "Pipeline_Full_Life_Cycle_Test_Dataset-v7:scenario_pipeformer_dispatch_019_session_001::turn_001": (
        "Keep candidate 2 first, but completely rank candidates 1 and 3 using their supported warning, pressure, linepack, and energy metrics."
    ),
}


EVALUATOR_FALSE_POSITIVES = {
    "Pipeline_Full_Life_Cycle_Test_Dataset-v4:scenario_pipeformer_dispatch_006_session_001::turn_001": "The two candidates have the same rounded linepack decline; the stated tie is defensible.",
    "Pipeline_Full_Life_Cycle_Test_Dataset-v4:scenario_pipeformer_dispatch_006_session_001::turn_002": "The two candidates have the same rounded linepack decline; the stated tie is defensible.",
    "Pipeline_Full_Life_Cycle_Test_Dataset-v4:scenario_pipeformer_dispatch_006_session_001::turn_003": "The values are grounded in prior turns; the current-turn numerical checker does not fully consume conversation evidence.",
    "Pipeline_Full_Life_Cycle_Test_Dataset-v4:scenario_pipeformer_dispatch_006_session_002::turn_001": "The pressure and energy comparison is supported by the prior candidate forecasts.",
    "Pipeline_Full_Life_Cycle_Test_Dataset-v4:scenario_pipeformer_prediction_018_session_002::turn_001": "The rejected numbers are derived arithmetic from 6%, 10%, and 20% thresholds.",
    "Pipeline_Full_Life_Cycle_Test_Dataset-v4:scenario_pipeformer_prediction_019_session_002::turn_001": "The 1- and 11-point margins are derived arithmetic from the 9%, 10%, and 20% values.",
    "Pipeline_Full_Life_Cycle_Test_Dataset-v7:scenario_pipeformer_dispatch_002_session_001::turn_001": "All candidates tie on violations, so selecting the energy leader follows the user's stated tiebreaker.",
    "Pipeline_Full_Life_Cycle_Test_Dataset-v7:scenario_pipeformer_dispatch_005_session_001::turn_001": "Candidate 2 is the only zero-warning linepack result and is also the energy leader; the raw decline difference is negligible.",
    "Pipeline_Full_Life_Cycle_Test_Dataset-v7:scenario_pipeformer_dispatch_007_session_001::turn_001": "All candidates pass with no warnings; choosing the energy leader is consistent with the available evidence.",
    "Pipeline_Full_Life_Cycle_Test_Dataset-v7:scenario_pipeformer_dispatch_010_session_001::turn_001": "Candidate 1 has the best supply-demand gap, which is the user's primary objective.",
    "Pipeline_Full_Life_Cycle_Test_Dataset-v7:scenario_pipeformer_dispatch_018_session_001::turn_001": "All candidates share a linepack warning; choosing the energy leader follows the user's energy-first objective.",
}


def main() -> int:
    records = _load_records()
    annotations = _load_annotations()
    evaluations = _load_evaluations()
    workbook = Workbook()
    workbook.remove(workbook.active)

    reviewed_ids = set(annotations)
    needs_ids = {
        sample_id
        for sample_id, value in annotations.items()
        if "Needs Review" in value["sheets"]
    }
    spot_ids = {
        sample_id
        for sample_id, value in annotations.items()
        if "Manual Spot Check" in value["sheets"]
    }
    failed_scenarios = {
        (record["dataset_source"], record["scenario_id"])
        for record in records
        if annotations.get(record["sample_id"], {}).get("final_disposition") == "failed"
    }
    new_uncovered = {
        key
        for key, value in ADDITIONAL_REGENERATION.items()
        if value[0] == "new_uncovered_scenario"
    }

    _write_summary(
        workbook,
        records,
        needs_ids,
        spot_ids,
        reviewed_ids,
        failed_scenarios,
        new_uncovered,
    )
    _write_regeneration(workbook, records, annotations)
    _write_answer_repairs(workbook, records, annotations)
    _write_false_positives(workbook, records, annotations, evaluations)
    _write_all_records(workbook, records, annotations, evaluations)
    polish_workbook(workbook, ILLEGAL_CHARACTERS_RE, summary_sheets={"Summary"})
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT)
    print(json.dumps(_verify(OUTPUT), ensure_ascii=False, indent=2))
    return 0


def _load_records() -> list[dict[str, Any]]:
    values = json.loads(MASTER_TRACE.read_text(encoding="utf-8-sig"))
    return [
        value
        for value in values
        if value.get("dataset_source") in PIPEFORMER_DATASETS
        and value.get("scenario_type") == "pipeformer"
    ]


def _load_annotations() -> dict[str, dict[str, Any]]:
    workbook = load_workbook(QUALITY_REPORT, read_only=True, data_only=False)
    result: dict[str, dict[str, Any]] = {}
    for sheet_name in ("Needs Review", "Manual Spot Check"):
        sheet = workbook[sheet_name]
        rows = sheet.iter_rows(values_only=True)
        headers = [str(value or "") for value in next(rows)]
        for values in rows:
            row = dict(zip(headers, values))
            if row.get("scenario_type") != "pipeformer":
                continue
            sample_id = str(row.get("sample_id") or "")
            if not sample_id:
                continue
            entry = result.setdefault(
                sample_id,
                {"sheets": [], "final_disposition": "", "reviewer_notes": []},
            )
            entry["sheets"].append(sheet_name)
            disposition = str(row.get("final_disposition") or "")
            if disposition:
                entry["final_disposition"] = disposition
            note = str(row.get("reviewer_notes") or "")
            if note and note not in entry["reviewer_notes"]:
                entry["reviewer_notes"].append(note)
    workbook.close()
    return result


def _load_evaluations() -> dict[str, dict[str, Any]]:
    if not AUDIT_EVALUATION.is_file():
        return {}
    return {
        value["sample_id"]: value
        for value in (
            json.loads(line)
            for line in AUDIT_EVALUATION.read_text(encoding="utf-8-sig").splitlines()
            if line.strip()
        )
    }


def _write_summary(
    workbook: Workbook,
    records: list[dict[str, Any]],
    needs_ids: set[str],
    spot_ids: set[str],
    reviewed_ids: set[str],
    failed_scenarios: set[tuple[str, str]],
    new_uncovered: set[tuple[str, str]],
) -> None:
    sheet = workbook.create_sheet("Summary")
    sheet.append(["PipeFormer Additional Quality Audit", "Result"])
    dataset_counts = Counter(record["dataset_source"] for record in records)
    existing_pipeformer_targets = {
        key for key in REGENERATION_TARGETS if key[0] in PIPEFORMER_DATASETS
    }
    missing_audited_targets = set(ADDITIONAL_REGENERATION) - set(REGENERATION_TARGETS)
    rows = [
        ("Scope", "PipeFormer scenario_type records in v4 and v7 only"),
        ("Total PipeFormer records", len(records)),
        ("v4 PipeFormer records", dataset_counts["Pipeline_Full_Life_Cycle_Test_Dataset-v4"]),
        ("v7 PipeFormer records", dataset_counts["Pipeline_Full_Life_Cycle_Test_Dataset-v7"]),
        ("Manual Spot Check rows", len(spot_ids)),
        ("Manual Spot Check rate", len(spot_ids) / len(records)),
        ("Automated Needs Review rows", len(needs_ids)),
        ("Needs Review / Spot Check overlap", len(needs_ids & spot_ids)),
        ("Unique records represented in workbook review sheets", len(reviewed_ids)),
        ("Unique workbook review coverage rate", len(reviewed_ids) / len(records)),
        ("Records outside both workbook review sheets", len({r['sample_id'] for r in records} - reviewed_ids)),
        ("Current planned PipeFormer regeneration scenarios", len(existing_pipeformer_targets)),
        ("Audited regeneration-list additions", len(ADDITIONAL_REGENERATION)),
        ("Audited additions missing from current plan", len(missing_audited_targets)),
        ("New scenarios not previously rejected", len(new_uncovered)),
        (
            "Previously rejected scenarios reclassified from answer repair",
            len(ADDITIONAL_REGENERATION) - len(new_uncovered),
        ),
        ("Additional deterministic answer edits", len(ANSWER_REPAIRS)),
        ("Documented evaluator false positives", len(EVALUATOR_FALSE_POSITIVES)),
        (
            "Decision",
            "All scenarios on the Regeneration sheet are now included in REGENERATION_TARGETS. Stage them and human-review every changed turn before merge.",
        ),
        (
            "Audit boundary",
            "This audit checks stored tool-call integrity, binary state handling, structured provenance, candidate ranking, user priorities, and answer/evidence consistency. It does not replace human sign-off on regenerated stochastic outputs.",
        ),
    ]
    for row in rows:
        sheet.append(row)


def _write_regeneration(
    workbook: Workbook,
    records: list[dict[str, Any]],
    annotations: dict[str, dict[str, Any]],
) -> None:
    sheet = workbook.create_sheet("Regeneration")
    headers = [
        "dataset_source",
        "scenario_id",
        "finding_class",
        "recommendation",
        "reason",
        "evidence_sample_id",
        "workbook_review_status",
        "scenario_record_count",
        "failed_tool_output_count",
        "binary_state_forecast_count",
    ]
    sheet.append(headers)
    grouped = _group_by_scenario(records)
    for key, (finding_class, reason, suffix) in sorted(ADDITIONAL_REGENERATION.items()):
        scenario_records = grouped[key]
        evidence_sample_id = f"{key[0]}:{suffix}"
        review_statuses = sorted(
            {
                annotations[record["sample_id"]]["final_disposition"]
                for record in scenario_records
                if record["sample_id"] in annotations
            }
        )
        sheet.append(
            [
                key[0],
                key[1],
                finding_class,
                "regenerate_complete_scenario",
                reason,
                evidence_sample_id,
                ", ".join(review_statuses) or "not_in_workbook_review_sheets",
                len(scenario_records),
                sum(_failed_output_count(record) for record in scenario_records),
                sum(_binary_state_call_count(record) for record in scenario_records),
            ]
        )


def _write_answer_repairs(
    workbook: Workbook,
    records: list[dict[str, Any]],
    annotations: dict[str, dict[str, Any]],
) -> None:
    sheet = workbook.create_sheet("Answer Repairs")
    sheet.append(
        [
            "sample_id",
            "dataset_source",
            "scenario_id",
            "recommendation",
            "repair_reason",
            "workbook_review_status",
            "user_input",
            "current_final_answer",
        ]
    )
    by_id = {record["sample_id"]: record for record in records}
    for sample_id, reason in ANSWER_REPAIRS.items():
        record = by_id[sample_id]
        annotation = annotations.get(sample_id, {})
        sheet.append(
            [
                sample_id,
                record["dataset_source"],
                record["scenario_id"],
                "deterministic_answer_edit",
                reason,
                annotation.get("final_disposition") or "not_in_workbook_review_sheets",
                record.get("user_input"),
                record.get("final_answer"),
            ]
        )


def _write_false_positives(
    workbook: Workbook,
    records: list[dict[str, Any]],
    annotations: dict[str, dict[str, Any]],
    evaluations: dict[str, dict[str, Any]],
) -> None:
    sheet = workbook.create_sheet("Evaluator False Positives")
    sheet.append(
        [
            "sample_id",
            "dataset_source",
            "scenario_id",
            "audit_disposition",
            "why_flag_is_not_a_regeneration_reason",
            "current_native_issues",
            "current_task1_failed_checks",
            "workbook_review_status",
        ]
    )
    by_id = {record["sample_id"]: record for record in records}
    for sample_id, reason in EVALUATOR_FALSE_POSITIVES.items():
        record = by_id[sample_id]
        evaluation = evaluations.get(sample_id, {})
        annotation = annotations.get(sample_id, {})
        sheet.append(
            [
                sample_id,
                record["dataset_source"],
                record["scenario_id"],
                "pass",
                reason,
                ", ".join(evaluation.get("native_quality_issues") or []),
                ", ".join(evaluation.get("task1_failed_checks") or []),
                annotation.get("final_disposition") or "not_in_workbook_review_sheets",
            ]
        )


def _write_all_records(
    workbook: Workbook,
    records: list[dict[str, Any]],
    annotations: dict[str, dict[str, Any]],
    evaluations: dict[str, dict[str, Any]],
) -> None:
    sheet = workbook.create_sheet("All PipeFormer Records")
    headers = [
        "sample_id",
        "dataset_source",
        "scenario_id",
        "session_id",
        "turn_id",
        "workbook_sheets",
        "workbook_disposition",
        "recommended_action",
        "failed_tool_output_count",
        "binary_state_forecast_count",
        "current_native_flag",
        "current_task1_flag",
        "current_native_issues",
        "current_task1_failed_checks",
        "reviewer_notes",
        "user_input",
        "final_answer",
    ]
    sheet.append(headers)
    existing_targets = set(REGENERATION_TARGETS)
    for record in records:
        sample_id = record["sample_id"]
        key = (record["dataset_source"], record["scenario_id"])
        annotation = annotations.get(sample_id, {})
        evaluation = evaluations.get(sample_id, {})
        if key in existing_targets:
            action = "existing_regeneration_target"
        elif key in ADDITIONAL_REGENERATION:
            action = "additional_regeneration_target"
        elif sample_id in ANSWER_REPAIRS:
            action = "additional_deterministic_answer_edit"
        elif sample_id in DETERMINISTIC_REPAIR_SAMPLE_IDS:
            action = "existing_deterministic_answer_repair"
        else:
            action = "no_additional_problem_found"
        sheet.append(
            [
                sample_id,
                record["dataset_source"],
                record["scenario_id"],
                record.get("session_id"),
                record.get("turn_id"),
                ", ".join(annotation.get("sheets") or []),
                annotation.get("final_disposition") or "not_reviewed",
                action,
                _failed_output_count(record),
                _binary_state_call_count(record),
                evaluation.get("native_quality_flag") or record.get("quality_flag"),
                evaluation.get("task1_quality_flag"),
                ", ".join(evaluation.get("native_quality_issues") or []),
                ", ".join(evaluation.get("task1_failed_checks") or []),
                "\n".join(annotation.get("reviewer_notes") or []),
                record.get("user_input"),
                record.get("final_answer"),
            ]
        )


def _group_by_scenario(
    records: Iterable[dict[str, Any]],
) -> dict[tuple[str, str], list[dict[str, Any]]]:
    result: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for record in records:
        result[(record["dataset_source"], record["scenario_id"])].append(record)
    return result


def _failed_output_count(record: dict[str, Any]) -> int:
    return sum(
        1
        for item in record.get("tool_outputs") or []
        if (item.get("output") or {}).get("error")
        or (item.get("output") or {}).get("success") is False
    )


def _binary_state_call_count(record: dict[str, Any]) -> int:
    return sum(
        1
        for call in record.get("tool_calls") or []
        if call.get("name") == "run_pipeformer_forecast"
        and str((call.get("arguments") or {}).get("disturbance_variable") or "").endswith(":ST")
    )


def _verify(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=True, data_only=False)
    formula_count = 0
    error_cells: list[str] = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formula_count += 1
                if isinstance(cell.value, str) and cell.value.startswith("#"):
                    error_cells.append(f"{sheet.title}!{cell.coordinate}")
    result = {
        "output": path.as_posix(),
        "sheet_count": len(workbook.sheetnames),
        "formula_count": formula_count,
        "error_count": len(error_cells),
        "error_cells": error_cells,
    }
    workbook.close()
    return result


if __name__ == "__main__":
    raise SystemExit(main())
